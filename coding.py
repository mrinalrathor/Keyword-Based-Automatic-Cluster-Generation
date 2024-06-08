import os
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.chart import BarChart, Reference


def html_to_text(html_content):
    """
    Parses HTML content and extracts plain text.

    Args:
        html_content (str): HTML content as a string.

    Returns:
        str: Extracted text content.
    """
    soup = BeautifulSoup(html_content, 'html.parser')

    # Extract text from <p> tags
    text = ' '.join(p.get_text(separator=' ') for p in soup.find_all('p'))

    # Basic cleaning (lowercase and remove stop words)
    stop_words = set(
        "a an and are as at be by for from has he in is it its of on that the to was were will with".split())
    cleaned_text = []
    for word in text.lower().split():
        if word not in stop_words and len(word) > 1:
            cleaned_text.append(word)
    return ' '.join(cleaned_text)

def write_to_excel(clusters, sheet_name, filename):
    """
    Writes data to an Excel spreadsheet and adds charts.

    Args:
        clusters (dict): Dictionary containing clusters with their terms.
        sheet_name (str): Name of the worksheet.
        filename (str): Name of the output Excel file.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    # Write headers
    headers = ["Cluster", "Term", "TF", "N-TF"]
    sheet.append(headers)

    row_offset = 2  # Start writing data from the second row
    
    for cluster, terms in clusters.items():
        for term, stats in terms.items():
            sheet.append([cluster, term, stats['tf'], stats['ntf']])

    # Create a bar chart for keyword-based clustering
    for cluster, terms in clusters.items():
        if terms:
            chart = BarChart()
            chart.title = f"Term Frequencies in '{cluster}' Cluster"
            chart.x_axis.title = "Terms"
            chart.y_axis.title = "Frequency"
            
            # Define the data range for the chart
            start_row = sheet.max_row - len(terms) + 1
            end_row = sheet.max_row
            data_ref = Reference(sheet, min_col=3, min_row=start_row, max_row=end_row)
            categories_ref = Reference(sheet, min_col=2, min_row=start_row, max_row=end_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories_ref)
            chart.shape = 4
            sheet.add_chart(chart, f"E{start_row}")

    wb.save(filename)

def analyze_text(text, output_filename, keywords):
    """
    Analyzes text to calculate TF/NTF and clusters terms based on keywords.

    Args:
        text (str): Text content to analyze.
        output_filename (str): Name of the output Excel file.
        keywords (list): List of keywords for clustering.
    """
    term_counts = {}
    words = text.lower().split()
    
    for word in words:
        term_counts[word] = term_counts.get(word, 0) + 1

    total_sq_tf = sum(v**2 for v in term_counts.values())
    
    # Initialize clusters
    clusters = {keyword: {} for keyword in keywords}
    clusters['others'] = {}  # Cluster for terms not matching any keyword

    # Assign terms to clusters
    for term, tf in term_counts.items():
        ntf = tf / (total_sq_tf**0.5)
        assigned = False
        for keyword in keywords:
            if keyword in term:  # Simple keyword matching
                clusters[keyword][term] = {'tf': tf, 'ntf': ntf}
                assigned = True
                break
        if not assigned:
            clusters['others'][term] = {'tf': tf, 'ntf': ntf}

    write_to_excel(clusters, "Sheet1", output_filename)

# Define your HTML content
html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Customer Reviews - Sports Shoes</title>
</head>
<body>
    <h2>Review 1</h2>
    <p>These shoes are amazing! They provide great comfort and support during long runs. 
        I highly recommend them for anyone looking for a good running shoe. 
        The size was perfect and the color looks even better in person. 
        A bit pricey, but definitely worth the investment.</p>

    <h2>Review 2</h2>
    <p>These shoes are a disappointment. The soles are very slippery, 
        which makes them unsafe for outdoor use. The material feels cheap 
        and they started to fall apart after just a few weeks. Not happy with 
        this purchase at all. I would not recommend them.</p>

    <h2>Review 3</h2>
    <p>These basketball shoes are fantastic! The ankle support is excellent 
        and they provide great traction on the court. I've been using them for 
        several months now and they're still holding up well. True to size 
        and very comfortable for playing for extended periods. A bit on the 
        heavy side, but the performance outweighs that.</p>

    <h2>Review 4</h2>
    <p>I bought these shoes for casual wear, but they're not very comfortable. 
        The inner sole is thin and offers no cushioning. The style is nice, 
        but they're not practical for everyday use. Disappointed with the 
        quality considering the price. Might be okay for occasional wear, 
        but not for walking long distances.</p>
</body>
</html>
"""

# Define your keywords for clustering
keywords = ['comfort', 'support', 'size', 'color', 'price', 'material', 'quality']

# Process the HTML content
text_content = html_to_text(html_content)
print(f"Extracted text content: {text_content[:500]}")  # Print first 500 characters of extracted text

# Output filename
output_filename = "analysis_output.xlsx"

# Analyze the text and generate the output Excel file with charts
analyze_text(text_content, output_filename, keywords)

print(f"Analysis completed and output file generated: {output_filename}")

analyze_text(text_content, output_filename, keywords)

print(f"Analysis completed and output file generated: {output_filename}")


